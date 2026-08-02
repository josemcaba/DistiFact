"""
Módulo que contiene la clase ProcesadorFacturas para procesar archivos de facturas.
Refactorizado para soportar múltiples desgloses de IVA por factura.
"""
import sys
import os
import logging
from importlib import import_module
from typing import List, Callable, Dict, Any, Optional

# Librerías de terceros
import pdfplumber
import fitz  # PyMuPDF
import openpyxl

# Módulos internos
from .extractor_imagenes import ExtractorImagenes
from .extractor_texto import ExtractorTexto
from modelo.factura import Factura
from modelo.empresa import Empresa

logger = logging.getLogger(__name__)

class ProcesadorFacturas:
    """
    Clase que procesa archivos de facturas en diferentes formatos.
    """
    def __init__(self):
        self._progreso_callback: Optional[Callable[[int, int], None]] = None
        self._mensaje_callback: Optional[Callable[[str, str], None]] = None
        self._factura_callback: Optional[Callable[[Dict[str, Any]], None]] = None
        self.extractor = ExtractorImagenes()
        self.ocr = ExtractorTexto()
    
    def set_callbacks(self, progreso_callback: Callable[[int, int], None], 
                     mensaje_callback: Callable[[str, str], None],
                     factura_callback: Optional[Callable[[Dict[str, Any]], None]] = None) -> None:
        """
        Establece las funciones de callback para reportar progreso y mensajes.
        
        Args:
            progreso_callback: Función para reportar progreso
            mensaje_callback: Función para mostrar mensajes
            factura_callback: Función para mostrar información de factura procesada
        """
        self._progreso_callback = progreso_callback
        self._mensaje_callback = mensaje_callback
        self._factura_callback = factura_callback  # <--- Añadir
    
    def _mostrar_mensaje(self, tipo: str, mensaje: str) -> None:
        """Muestra un mensaje usando el callback si está disponible."""
        if self._mensaje_callback:
            self._mensaje_callback(tipo, mensaje)
        else:
            logger.info("%s: %s", tipo.upper(), mensaje)
    
    def _actualizar_progreso(self, actual: int, total: int) -> None:
        """Actualiza el progreso usando el callback si está disponible."""
        if self._progreso_callback:
            self._progreso_callback(actual, total)
        else:
            logger.info("Procesando: %d/%d", actual, total)
    
    def _mostrar_info_factura(self, info_factura: Dict[str, Any]) -> None:
        """Muestra información de una factura usando el callback """
        if self._factura_callback:
            self._factura_callback(info_factura)

    def _cargar_modulo_extractor(self, nombre_funcion: str):
        """Intenta cargar dinámicamente el módulo extractor."""
        try:
            # Quitamos extensión .py si existe y construimos ruta de paquete
            nombre_modulo = nombre_funcion.replace('.py', '')
            modulo_path = f"extractores.{nombre_modulo}"
            return import_module(modulo_path)
        except ImportError as e:
            self._mostrar_mensaje('error', f'No se pudo cargar el módulo "{nombre_funcion}": {e}')
            return None

    def _generar_objetos_factura(self, resultado_extractor: Any, num_pagina: int) -> List[Factura]:
        """
        Normaliza la salida del extractor y genera una lista de objetos Factura.
        Soporta:
        1. Lista de diccionarios (Nuevo estándar: múltiples IVAs).
           Ejemplo: [{...IVA1...}, {...IVA2...}]
        2. Lista [num_pag, dict] (Estándar antiguo).
        3. Dicccionario único (Caso simple).
        """
        facturas_generadas = []

        if not resultado_extractor:
            return []

        # Caso 1: Nuevo Estándar - Lista de diccionarios (Uno por tipo de IVA)
        if isinstance(resultado_extractor, list) and len(resultado_extractor) > 0:
            if isinstance(resultado_extractor[0], dict):
                # Lista pura de diccionarios
                for datos in resultado_extractor:
                    factura = Factura(num_pagina, datos)
                    facturas_generadas.append(factura)

            # Caso 2: Estándar Antiguo - [num_pag, dict]
            elif len(resultado_extractor) == 2 and isinstance(resultado_extractor[1], dict):
                datos = resultado_extractor[1]
                factura = Factura(num_pagina, datos)
                facturas_generadas.append(factura)

            else:
                self._mostrar_mensaje('error', f'Formato de lista no reconocido en página {num_pagina}')

        # Caso 3: Diccionario único
        elif isinstance(resultado_extractor, dict):
            factura = Factura(num_pagina, resultado_extractor)
            facturas_generadas.append(factura)

        else:
            self._mostrar_mensaje('error', f'Formato de retorno desconocido en página {num_pagina}: {type(resultado_extractor)}')
            return []

        # Mostrar información de la última factura procesada
        if facturas_generadas:
            self._mostrar_info_factura(facturas_generadas[-1].to_dict())

        return facturas_generadas
    
    def procesar_archivo(self, ruta_archivo: str, empresa: Empresa) -> List[Factura]:
        """
        Procesa un archivo según el tipo de empresa y genera las facturas.
        """
        if not os.path.exists(ruta_archivo):
            self._mostrar_mensaje('error', f'El archivo "{ruta_archivo}" no existe.')
            return []
        
        # 1. Cargar Módulo Extractor
        fe = self._cargar_modulo_extractor(empresa.funciones)
        if not fe:
            return []
        
        # 2. Obtener contenido crudo (texto por páginas/filas)
        paginas_datos = []
        if empresa.tipo == "PDFtexto":
            paginas_datos = self._procesar_pdf_texto(ruta_archivo, fe.identificador)
        elif empresa.tipo == "PDFimagen":
            paginas_datos = self._procesar_pdf_imagen(ruta_archivo, fe.identificador, empresa.nif)
        elif empresa.tipo == "excel":
            paginas_datos = self._procesar_excel(ruta_archivo, fe.identificador, empresa.nif)
        else:
            self._mostrar_mensaje('error', f'Tipo de archivo "{empresa.tipo}" no válido')
            return []
        
        if not paginas_datos:
            return []
        
        # 3. Extraer datos estructurados y crear objetos Factura
        facturas_finales = []
        
        for pagina in paginas_datos:
            try:
                # pagina es [num_pagina, contenido]
                num_pagina = pagina[0]
                
                # Llamada al extractor específico
                # IMPORTANTE: Se pasa pagina completa porque algunos extractores viejos usan pagina[0]
                datos_crudos = fe.extraerDatosFactura(pagina, empresa.to_dict())
                
                # Normalizar y convertir a objetos Factura
                nuevas_facturas = self._generar_objetos_factura(datos_crudos, num_pagina)
                facturas_finales.extend(nuevas_facturas)

            except Exception as e:
                self._mostrar_mensaje('error', f'Excepción procesando página {pagina[0]}: {str(e)}')
                import traceback
                traceback.print_exc()

        if not facturas_finales:
            self._mostrar_mensaje('warning', f'El archivo se leyó pero no se extrajeron facturas válidas.')
        
        return facturas_finales
    
    # -------------------------------------------------------------------------
    # Métodos de extracción de contenido por tipo de archivo (Sin cambios mayores)
    # -------------------------------------------------------------------------

    def _procesar_pdf_texto(self, ruta_pdf: str, identificador: str) -> List[List]:
        paginas = []
        paginas_descartadas = []
        try:
            with pdfplumber.open(ruta_pdf) as pdf:
                total = len(pdf.pages)
                for i, pagina in enumerate(pdf.pages, start=1):
                    self._actualizar_progreso(i, total)
                    texto = pagina.extract_text() or ""
                    
                    if identificador in texto:
                        paginas.append([i, texto])
                    else:
                        paginas_descartadas.append(str(i))
            
            self._mostrar_mensaje('info', f'Procesadas {total} páginas del archivo PDF')
            self._mostrar_mensaje('info', f'Encontradas {len(paginas)} páginas válidas')
            if paginas_descartadas:
                self._mostrar_mensaje('info', f"Páginas descartadas: {', '.join(paginas_descartadas)}")
            return paginas
        except Exception as e:
            self._mostrar_mensaje('error', f'Error lectura PDF Texto: {e}')
            return []

    def _procesar_pdf_imagen(self, ruta_pdf: str, identificador: str, nif: str) -> List[List]:
        rectangulos = self.extractor.cargar_rectangulos_json(nif)
        if not rectangulos:
            self._mostrar_mensaje('error', f'Faltan rectángulos para NIF {nif}')
            return []
        
        angulo = rectangulos.get("angulo", 0)
        paginas = []
        paginas_descartadas = []
        try:
            with fitz.open(ruta_pdf) as doc:
                total = doc.page_count
                for i in range(total):
                    self._actualizar_progreso(i + 1, total)
                    
                    # Extracción visual y OCR
                    img, angulo = self.extractor.extraer_imagen_de_pdf(doc, i, angulo)
                    imgs_recorte = self.extractor.extraer_imagenes_de_rectangulos(img, rectangulos)
                    texto = self.ocr.extraer_texto_de_las_imagenes(imgs_recorte)
                    
                    if texto and identificador in texto:
                        paginas.append([i + 1, texto])
                    else:
                        paginas_descartadas.append(str(i + 1))
            
            self._mostrar_mensaje('info', f'Procesadas {total} páginas en archivo PDF')
            self._mostrar_mensaje('info', f'Encontradas {len(paginas)} páginas válidas')
            if paginas_descartadas:
                self._mostrar_mensaje('info', f"Páginas descartadas: {' - '.join(paginas_descartadas)}")
            return paginas
        except Exception as e:
            self._mostrar_mensaje('error', f'Error lectura PDF Imagen: {e}')
            return []
    
    def _procesar_excel(self, ruta_excel: str, identificador: str, nif: str) -> List[List]:
        filas_procesadas = []
        try:
            libro = openpyxl.load_workbook(ruta_excel, data_only=True)
            hoja = libro.active
            max_row = hoja.max_row
            
            # Iteramos filas. openpyxl es base-1, start=2 para saltar cabecera
            for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                self._actualizar_progreso(i, max_row)
                
                # Si la fila tiene algún dato
                if any(c is not None for c in fila):
                    # Formato: [NumeroFila, ListaValores]
                    # Convertimos tupla a lista para consistencia
                    filas_procesadas.append([i, list(fila)])
            
            self._mostrar_mensaje('info', f'Procesadas {len(filas_procesadas)} filas de Excel.')
            return filas_procesadas
            
        except Exception as e:
            self._mostrar_mensaje('error', f'Error lectura Excel: {e}')
            return []